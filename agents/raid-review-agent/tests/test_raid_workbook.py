import hashlib
import importlib.util
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook


MODULE_PATH = Path(__file__).parents[1] / "scripts" / "raid_workbook.py"
SPEC = importlib.util.spec_from_file_location("raid_workbook", MODULE_PATH)
raid_workbook = importlib.util.module_from_spec(SPEC)
assert SPEC.loader
SPEC.loader.exec_module(raid_workbook)


def write_json(path: Path, document: dict) -> None:
    path.write_text(json.dumps(document, indent=2), encoding="utf-8")


def test_sanitize_template_clears_project_data_and_corrects_decision_header(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "sanitized.xlsx"
    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover"
    cover["A5"] = "Example data - remove before use"
    plan = workbook.create_sheet("Project_Plan")
    plan.append(["#", "Task"])
    plan.append([1, "Private project task"])
    risks = workbook.create_sheet("Risk Tracker")
    risks.append(["Date", "Risk"])
    risks.append(["Ongoing", "Private risk"])
    actions = workbook.create_sheet("Action Items")
    actions.append(["ID", "Action Item"])
    actions.append([1, "Private action"])
    decisions = workbook.create_sheet("Decision Log")
    decisions.append(["t", "Decision"])
    decisions.append([1, "Private decision"])
    workbook.save(source)

    raid_workbook.sanitize_template(source, destination)

    sanitized = load_workbook(destination)
    assert sanitized["Project_Plan"]["B2"].value is None
    assert sanitized["Risk Tracker"]["B2"].value is None
    assert sanitized["Action Items"]["B2"].value is None
    assert sanitized["Decision Log"]["B2"].value is None
    assert sanitized["Decision Log"]["A1"].value == "ID"
    assert "no project data" in sanitized["Cover"]["A5"].value


def test_apply_approved_creates_backup_output_and_audit(tmp_path):
    workbook_path = tmp_path / "raid.xlsx"
    workbook = Workbook()
    risks = workbook.active
    risks.title = "Risk Tracker"
    risks.append(["ID", "Risk", "Notes", "Status"])
    risks.append(["R-1", "Schedule risk", "Initial note", "Active"])
    workbook.save(workbook_path)

    proposal_path = tmp_path / "proposal.json"
    approval_path = tmp_path / "approval.json"
    output_path = tmp_path / "updated.xlsx"
    audit_path = tmp_path / "audit.json"
    workbook_hash = raid_workbook.sha256_file(workbook_path)
    proposal = {
        "version": 1,
        "workbook_path": str(workbook_path),
        "workbook_sha256": workbook_hash,
        "created_at": "2026-08-20T12:00:00-07:00",
        "changes": [
            {
                "id": "P-001",
                "operation": "update",
                "sheet": "Risk Tracker",
                "match": {"field": "ID", "value": "R-1"},
                "values": {"Notes": {"value": "[8/20 update] Mitigation assigned.", "mode": "append"}},
                "rationale": "Current project evidence identifies an owner.",
                "sources": ["Meeting Monitor: 2026-08-20 project sync"],
                "confidence": "high"
            },
            {
                "id": "P-002",
                "operation": "insert",
                "sheet": "Risk Tracker",
                "key_field": "ID",
                "values": {
                    "ID": {"value": "R-2"},
                    "Risk": {"value": "Dependency risk"},
                    "Status": {"value": "Active"}
                },
                "rationale": "A new dependency was confirmed.",
                "sources": ["Status report: 2026-08-20"],
                "confidence": "medium"
            }
        ]
    }
    write_json(proposal_path, proposal)
    approval = {
        "version": 1,
        "proposal_sha256": hashlib.sha256(proposal_path.read_bytes()).hexdigest(),
        "workbook_sha256": workbook_hash,
        "approved_at": "2026-08-20T12:05:00-07:00",
        "items": [
            {"proposal_id": "P-001", "disposition": "approved"},
            {"proposal_id": "P-002", "disposition": "approved"}
        ]
    }
    write_json(approval_path, approval)

    raid_workbook.apply_approved(
        workbook_path,
        proposal_path,
        approval_path,
        output_path,
        audit_path,
        update_current=False,
        confirmation=None,
    )

    result = load_workbook(output_path)["Risk Tracker"]
    assert result["C2"].value == "Initial note [8/20 update] Mitigation assigned."
    assert result["A3"].value == "R-2"
    assert list(tmp_path.glob("raid.backup_*.xlsx"))
    audit = json.loads(audit_path.read_text(encoding="utf-8"))
    assert audit["validation"] == "passed"
    assert len(audit["applied"]) == 2


def test_uncertain_proposal_cannot_be_written(tmp_path):
    workbook_path = tmp_path / "raid.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Risk Tracker"
    worksheet.append(["ID", "Risk"])
    worksheet.append(["R-1", "Risk"])
    workbook.save(workbook_path)
    workbook_hash = raid_workbook.sha256_file(workbook_path)
    proposal_path = tmp_path / "proposal.json"
    proposal = {
        "version": 1,
        "workbook_path": str(workbook_path),
        "workbook_sha256": workbook_hash,
        "created_at": "2026-08-20T12:00:00-07:00",
        "changes": [{
            "id": "P-001",
            "operation": "update",
            "sheet": "Risk Tracker",
            "match": {"field": "ID", "value": "R-1"},
            "values": {"Risk": {"value": "Guess"}},
            "rationale": "Unverified report.",
            "sources": ["Unverified note"],
            "confidence": "uncertain"
        }]
    }
    write_json(proposal_path, proposal)
    approval_path = tmp_path / "approval.json"
    write_json(approval_path, {
        "version": 1,
        "proposal_sha256": hashlib.sha256(proposal_path.read_bytes()).hexdigest(),
        "workbook_sha256": workbook_hash,
        "approved_at": "2026-08-20T12:05:00-07:00",
        "items": [{"proposal_id": "P-001", "disposition": "approved"}]
    })

    try:
        raid_workbook.apply_approved(
            workbook_path,
            proposal_path,
            approval_path,
            tmp_path / "out.xlsx",
            tmp_path / "audit.json",
            update_current=False,
            confirmation=None,
        )
    except ValueError as error:
        assert "Uncertain proposal" in str(error)
    else:
        raise AssertionError("Uncertain proposal was incorrectly written")