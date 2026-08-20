from __future__ import annotations

import argparse
import copy
import hashlib
import json
import shutil
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from jsonschema import Draft202012Validator
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


PACKAGE_ROOT = Path(__file__).resolve().parents[1]
SCHEMA_ROOT = PACKAGE_ROOT / "schemas"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def load_json(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as stream:
        return json.load(stream)


def validate_document(document: dict[str, Any], schema_name: str) -> None:
    schema = load_json(SCHEMA_ROOT / schema_name)
    errors = sorted(Draft202012Validator(schema).iter_errors(document), key=lambda error: list(error.path))
    if errors:
        details = "\n".join(f"- {'/'.join(map(str, error.path)) or '<root>'}: {error.message}" for error in errors)
        raise ValueError(f"Schema validation failed:\n{details}")


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def workbook_summary(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    sheets = []
    for worksheet in workbook.worksheets:
        headers = [json_value(worksheet.cell(1, column).value) for column in range(1, worksheet.max_column + 1)]
        sheets.append(
            {
                "name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "headers": headers,
                "freeze_panes": str(worksheet.freeze_panes) if worksheet.freeze_panes else None,
                "auto_filter": worksheet.auto_filter.ref,
                "tables": [{"name": name, "ref": worksheet.tables[name].ref} for name in worksheet.tables],
                "formula_count": sum(
                    1 for row in worksheet.iter_rows() for cell in row if not isinstance(cell, MergedCell) and cell.data_type == "f"
                ),
                "data_validation_count": len(worksheet.data_validations.dataValidation),
                "protected": bool(worksheet.protection.sheet),
            }
        )
    return {"path": str(path.resolve()), "sha256": sha256_file(path), "sheets": sheets}


def normalize_text(value: str) -> str:
    return value.replace("\u2014", "-").replace("\u2013", "-").replace("\u2011", "-").replace("\u2192", "to")


def clear_values(worksheet, start_row: int) -> None:
    for row in worksheet.iter_rows(min_row=start_row):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def sanitize_template(source: Path, destination: Path) -> None:
    if source.resolve() == destination.resolve():
        raise ValueError("The sanitized destination must differ from the source template.")
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    workbook = load_workbook(destination, data_only=False)

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell, MergedCell) and isinstance(cell.value, str):
                    cell.value = normalize_text(cell.value)

    if "Project_Plan" in workbook.sheetnames:
        clear_values(workbook["Project_Plan"], 2)
    if "Risk Tracker" in workbook.sheetnames:
        clear_values(workbook["Risk Tracker"], 2)
    if "Action Items" in workbook.sheetnames:
        clear_values(workbook["Action Items"], 2)
    if "Decision Log" in workbook.sheetnames:
        clear_values(workbook["Decision Log"], 2)
        workbook["Decision Log"]["A1"] = "ID"
    if "Cover" in workbook.sheetnames:
        workbook["Cover"]["A5"] = (
            "This workbook contains a Project Plan tab and RAID tabs for risks, action items, and decisions.\n\n"
            "The included template contains no project data. Configure the RAID Review Agent before use."
        )

    workbook.save(destination)


def header_map(worksheet) -> dict[str, int]:
    return {
        str(worksheet.cell(1, column).value).strip(): column
        for column in range(1, worksheet.max_column + 1)
        if worksheet.cell(1, column).value is not None
    }


def find_row(worksheet, change: dict[str, Any]) -> int:
    if change.get("row"):
        return int(change["row"])
    match = change.get("match")
    if not match:
        raise ValueError(f"Change {change['id']} requires row or match.")
    columns = header_map(worksheet)
    if match["field"] not in columns:
        raise ValueError(f"Match field {match['field']!r} is not present on {worksheet.title!r}.")
    column = columns[match["field"]]
    matches = [row for row in range(2, worksheet.max_row + 1) if worksheet.cell(row, column).value == match["value"]]
    if len(matches) != 1:
        raise ValueError(f"Change {change['id']} expected one row match but found {len(matches)}.")
    return matches[0]


def first_blank_row(worksheet, key_column: int) -> int:
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row, key_column).value in (None, ""):
            return row
    return worksheet.max_row + 1


def copy_row_style(worksheet, source_row: int, target_row: int) -> None:
    if source_row < 1 or source_row == target_row:
        return
    for column in range(1, worksheet.max_column + 1):
        source = worksheet.cell(source_row, column)
        target = worksheet.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy.copy(source.alignment)
        target.protection = copy.copy(source.protection)
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height


def update_table_refs(worksheet, row: int) -> None:
    for name in worksheet.tables:
        table = worksheet.tables[name]
        start, end = table.ref.split(":")
        end_column = "".join(character for character in end if character.isalpha())
        end_row = int("".join(character for character in end if character.isdigit()))
        if row > end_row:
            table.ref = f"{start}:{end_column}{row}"


def apply_change(workbook, change: dict[str, Any]) -> dict[str, Any]:
    worksheet = workbook[change["sheet"]]
    columns = header_map(worksheet)
    operation = change["operation"]
    if operation == "update":
        row = find_row(worksheet, change)
    else:
        key_field = change.get("key_field") or next(iter(change["values"]))
        if key_field not in columns:
            raise ValueError(f"Key field {key_field!r} is not present on {worksheet.title!r}.")
        row = first_blank_row(worksheet, columns[key_field])
        if row > worksheet.max_row:
            copy_row_style(worksheet, max(2, row - 1), row)

    before: dict[str, Any] = {}
    after: dict[str, Any] = {}
    for field, instruction in change["values"].items():
        if field not in columns:
            raise ValueError(f"Field {field!r} is not present on {worksheet.title!r}.")
        cell = worksheet.cell(row, columns[field])
        before[field] = json_value(cell.value)
        value = instruction["value"]
        value_type = instruction.get("value_type", "string")
        if value is not None and value_type == "date":
            value = date.fromisoformat(str(value))
        elif value is not None and value_type == "integer":
            value = int(value)
        elif value is not None and value_type == "number":
            value = float(value)
        if instruction.get("mode", "replace") == "append" and cell.value not in (None, ""):
            separator = instruction.get("separator", " ")
            value = f"{cell.value}{separator}{value}"
        cell.value = value
        after[field] = json_value(cell.value)

    if worksheet.auto_filter.ref:
        end_column = get_column_letter(worksheet.max_column)
        worksheet.auto_filter.ref = f"A1:{end_column}{max(worksheet.max_row, row)}"
    update_table_refs(worksheet, row)
    return {"proposal_id": change["id"], "sheet": worksheet.title, "row": row, "before": before, "after": after}


def apply_approved(
    workbook_path: Path,
    proposal_path: Path,
    approval_path: Path,
    output_path: Path,
    audit_path: Path,
    update_current: bool,
    confirmation: str | None,
) -> None:
    proposal = load_json(proposal_path)
    approval = load_json(approval_path)
    validate_document(proposal, "raid-proposal.schema.json")
    validate_document(approval, "raid-approval.schema.json")

    current_hash = sha256_file(workbook_path)
    if current_hash != proposal["workbook_sha256"] or current_hash != approval["workbook_sha256"]:
        raise ValueError("Workbook fingerprint changed after review. Run a new review before applying changes.")
    if sha256_file(proposal_path) != approval["proposal_sha256"]:
        raise ValueError("Approval does not match the supplied proposal file.")
    if update_current and confirmation != str(workbook_path.resolve()):
        raise ValueError("Current-workbook mode requires the exact resolved workbook path as confirmation.")

    known_ids = {change["id"] for change in proposal["changes"]}
    approval_ids = {item["proposal_id"] for item in approval["items"]}
    if approval_ids != known_ids:
        missing = sorted(known_ids - approval_ids)
        unknown = sorted(approval_ids - known_ids)
        raise ValueError(f"Every proposal requires one disposition. Missing={missing}, unknown={unknown}")
    if len(approval_ids) != len(approval["items"]):
        raise ValueError("Each proposal may appear only once in the approval manifest.")

    approvals = {item["proposal_id"]: item for item in approval["items"]}
    changes = []
    for original in proposal["changes"]:
        approval_item = approvals[original["id"]]
        if approval_item["disposition"] != "approved":
            continue
        if original["confidence"] == "uncertain":
            raise ValueError(f"Uncertain proposal {original['id']} cannot be approved for writing.")
        change = copy.deepcopy(original)
        if approval_item.get("overrides"):
            change["values"].update(approval_item["overrides"])
        changes.append(change)
    if not approval_ids <= known_ids:
        raise ValueError("Approval references a proposal item that does not exist.")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = workbook_path.with_name(f"{workbook_path.stem}.backup_{timestamp}{workbook_path.suffix}")
    shutil.copy2(workbook_path, backup)
    target = workbook_path if update_current else output_path
    if not update_current and target.exists():
        raise FileExistsError(f"Output already exists: {target}")

    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(workbook_path, data_only=False, keep_vba=keep_vba)
    applied = [apply_change(workbook, change) for change in changes]
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)

    verified = load_workbook(target, data_only=False, keep_vba=keep_vba)
    for record in applied:
        worksheet = verified[record["sheet"]]
        columns = header_map(worksheet)
        for field, expected in record["after"].items():
            actual = json_value(worksheet.cell(record["row"], columns[field]).value)
            if actual != expected:
                raise ValueError(f"Post-write validation failed for {record['proposal_id']} field {field}.")

    audit = {
        "applied_at": datetime.now().astimezone().isoformat(),
        "source_workbook": str(workbook_path.resolve()),
        "output_workbook": str(target.resolve()),
        "backup_workbook": str(backup.resolve()),
        "proposal": str(proposal_path.resolve()),
        "approval": str(approval_path.resolve()),
        "applied": applied,
        "validation": "passed",
    }
    audit_path.parent.mkdir(parents=True, exist_ok=True)
    audit_path.write_text(json.dumps(audit, indent=2, ensure_ascii=True), encoding="utf-8")
    print(json.dumps(audit, indent=2, ensure_ascii=True))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="RAID workbook inspection and approval-gated update tool")
    subparsers = parser.add_subparsers(dest="command", required=True)

    inspect_parser = subparsers.add_parser("inspect")
    inspect_parser.add_argument("workbook", type=Path)

    fingerprint_parser = subparsers.add_parser("fingerprint")
    fingerprint_parser.add_argument("workbook", type=Path)

    sanitize_parser = subparsers.add_parser("sanitize-template")
    sanitize_parser.add_argument("source", type=Path)
    sanitize_parser.add_argument("destination", type=Path)

    copy_parser = subparsers.add_parser("copy-template")
    copy_parser.add_argument("source", type=Path)
    copy_parser.add_argument("destination", type=Path)

    validate_parser = subparsers.add_parser("validate-json")
    validate_parser.add_argument("schema")
    validate_parser.add_argument("document", type=Path)

    apply_parser = subparsers.add_parser("apply")
    apply_parser.add_argument("--workbook", required=True, type=Path)
    apply_parser.add_argument("--proposal", required=True, type=Path)
    apply_parser.add_argument("--approval", required=True, type=Path)
    apply_parser.add_argument("--output", required=True, type=Path)
    apply_parser.add_argument("--audit", required=True, type=Path)
    apply_parser.add_argument("--update-current", action="store_true")
    apply_parser.add_argument("--confirm-current")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        if args.command == "inspect":
            print(json.dumps(workbook_summary(args.workbook), indent=2, ensure_ascii=True))
        elif args.command == "fingerprint":
            print(sha256_file(args.workbook))
        elif args.command == "sanitize-template":
            sanitize_template(args.source, args.destination)
            print(json.dumps(workbook_summary(args.destination), indent=2, ensure_ascii=True))
        elif args.command == "copy-template":
            if args.destination.exists():
                raise FileExistsError(f"Destination already exists: {args.destination}")
            args.destination.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(args.source, args.destination)
            print(json.dumps(workbook_summary(args.destination), indent=2, ensure_ascii=True))
        elif args.command == "validate-json":
            validate_document(load_json(args.document), args.schema)
            print(f"VALID: {args.document}")
        elif args.command == "apply":
            apply_approved(
                args.workbook,
                args.proposal,
                args.approval,
                args.output,
                args.audit,
                args.update_current,
                args.confirm_current,
            )
        return 0
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())